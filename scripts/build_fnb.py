# -*- coding: utf-8 -*-
"""
build_fnb.py — F&B Службы питания Утрау (ресторан КАЕФ).

НОВАЯ схема (2026-07): каждый месяц = отдельный Google-файл (выгрузка iiko OLAP),
ссылки собираются в РЕЕСТРЕ-таблице (месяц -> ссылка на файл).
Парсер обходит ВСЕ вкладки книги, классифицирует их по СТРУКТУРЕ
(продажи / акты списания), собирает сводку -> data/fnb_data.js (window.UTRAU_FNB).

Реестр: 1ueE5blgChEYljlVnGktWKedsrcxfXXcn-U64X80AOQA
  колонка A = месяц (напр. «Июль 2026»), B = ссылка на файл месяца.
  Необязательные колонки (по ЗАГОЛОВКУ, если добавлены строкой сверху):
    «аренда зала», «гости а-ля карт», «гости банкет», «товарный остаток»,
    «пробковый», «услуги» — то, чего нет в выгрузке iiko.

Новый месяц = вставить строку (месяц + ссылка) в реестр. Файл должен быть
открыт «Доступ по ссылке -> Просмотр». Ничего в коде править не нужно.

Запуск: python scripts/build_fnb.py
"""
import re, io, json, csv, sys, urllib.request, datetime
from pathlib import Path
import openpyxl

try: sys.stdout.reconfigure(encoding="utf-8")   # Windows-консоль иначе падает на эмодзи/кириллице
except Exception: pass

ROOT = Path(__file__).resolve().parent.parent
OUT  = ROOT / "data" / "fnb_data.js"
RAW  = ROOT / "data" / "fnb_raw"
RAW.mkdir(parents=True, exist_ok=True)

REGISTRY_ID = "1ueE5blgChEYljlVnGktWKedsrcxfXXcn-U64X80AOQA"

MONTHS_RU = {"янв":1,"фев":2,"мар":3,"апр":4,"май":5,"мая":5,"июн":6,
             "июл":7,"авг":8,"сен":9,"окт":10,"ноя":11,"дек":12}

# ── утилиты ───────────────────────────────────────────────────────────────
def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return round(float(v), 2)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(" ", "")
    s = s.replace("₽", "").replace("%", "").replace(",", ".")
    if s in ("", "-", "—", "."): return None
    try: return round(float(s), 2)
    except ValueError: return None

def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def month_key(text):
    s = norm(text)
    mon = None
    for k, v in MONTHS_RU.items():
        if s.startswith(k):
            mon = v; break
    ym = re.search(r"(20\d{2})", s)
    if mon and ym:
        return f"{ym.group(1)}-{mon:02d}"
    return None

def file_id(url):
    m = re.search(r"/spreadsheets/d/([A-Za-z0-9_\-]+)", str(url or ""))
    if m: return m.group(1)
    m = re.search(r"[?&]id=([A-Za-z0-9_\-]+)", str(url or ""))
    return m.group(1) if m else None

def http_get(url, timeout=60):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()

# ── реестр ────────────────────────────────────────────────────────────────
# Необязательные показатели: ключ summary -> ключевые слова в заголовке колонки
EXTRA_COLS = {
    "rev_banquet_hall": ["аренда", "зал"],
    "rev_banquet_cork": ["пробков", "cork"],
    "rev_services_total": ["услуг"],
    "guests_alacarte": ["гост", "карт"],
    "guests_banquet": ["гост", "банкет"],
    "inv_total": ["остаток", "инвентар", "товарн"],
}

def match_extra(header_cell):
    h = norm(header_cell)
    if not h: return None
    for key, words in EXTRA_COLS.items():
        if all(w in h for w in words):
            return key
    return None

def fetch_registry():
    """-> список dict: {month, file_id, extras{}}"""
    url = f"https://docs.google.com/spreadsheets/d/{REGISTRY_ID}/export?format=csv"
    raw = http_get(url, 30).decode("utf-8", errors="replace")
    if raw.lstrip().startswith("<"):
        raise RuntimeError("реестр недоступен (нет доступа по ссылке?)")
    rows = list(csv.reader(io.StringIO(raw)))

    # заголовок? (первая непустая строка, где A не распознаётся как месяц)
    extra_map = {}   # индекс колонки -> ключ summary
    start = 0
    for i, r in enumerate(rows):
        if not any(str(c).strip() for c in r):
            continue
        if month_key(r[0]) is None:            # это строка-заголовок
            for j, c in enumerate(r):
                k = match_extra(c)
                if k: extra_map[j] = k
            start = i + 1
        break

    out = []
    for r in rows[start:]:
        if not r or not str(r[0]).strip():
            continue
        mk = month_key(r[0])
        fid = file_id(r[1]) if len(r) > 1 else None
        if not mk or not fid:
            continue
        extras = {}
        for j, key in extra_map.items():
            if j < len(r):
                v = num(r[j])
                if v is not None:
                    extras[key] = v
        out.append({"month": mk, "file_id": fid, "extras": extras})
    return out

# ── классификация вкладок ───────────────────────────────────────────────────
SALE_CATS = [   # (ключевые слова в имени вкладки -> категория продаж)
    (["a-la carte", "a la carte", "а-ля карт", "меню a"], "kitchen"),
    (["бар напит", "напитки"], "bar_soft"),
    (["бар алког", "алкогол"], "bar_alco"),
    (["мероприят", "банкет"], "banquet"),
    (["завтрак"], "breakfast"),
    (["без оплат"], "free"),
]
WO_CATS = [     # (ключевые слова -> категория списаний)
    (["удален"], "wo_deletions"),
    (["комплемент", "комплимент"], "wo_comp"),
    (["представит"], "wo_represent"),
    (["стафф", "питание сотруд", "сотрудник"], "wo_staff"),
    (["порча"], "wo_spoilage"),
    (["проработ"], "wo_dev"),
]

def cat_by_name(name, table):
    n = norm(name)
    for words, cat in table:
        if any(w in n for w in words):
            return cat
    return None

def cat_by_content(rows, table):
    """Категория списания по колонке «Счёт списания» (когда имя вкладки — мусор, напр. «Лист1»)."""
    ci = -1
    for r in rows:
        for j, c in enumerate(r):
            if norm(c) in ("счёт списания", "счет списания"):
                ci = j; break
        if ci >= 0: break
    if ci < 0: return None
    txt = " ".join(norm(r[ci]) for r in rows if ci < len(r) and r[ci])
    for words, cat in table:
        if any(w in txt for w in words):
            return cat
    return None

DATE_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4}")

def find_header(rows, *needles):
    needles = [n.lower() for n in needles]
    for i, r in enumerate(rows):
        joined = norm(" | ".join("" if c is None else str(c) for c in r))
        if all(n in joined for n in needles):
            return i
    return -1

def classify(name, rows):
    """-> ('sale'|'writeoff'|'unknown', hint)"""
    if find_header(rows, "блюдо", "сумма со скидкой") >= 0:
        return "sale", cat_by_name(name, SALE_CATS)
    # акты списания: где-то «акты списания» или колонка «сумма, р.» + строки-даты
    has_acts = any("акты списания" in norm(" ".join("" if c is None else str(c) for c in r)) for r in rows[:3])
    has_sum  = find_header(rows, "сумма, р") >= 0
    has_dates = any(rows and DATE_RE.match(str(r[0]).strip()) for r in rows if r and r[0] is not None)
    if has_acts or (has_sum and has_dates):
        return "writeoff", cat_by_name(name, WO_CATS)
    return "unknown", None

def period_month(rows):
    """месяц из строки 'Период: с 01.07.2026 ...' внутри листа -> 'YYYY-MM' | None"""
    for r in rows[:6]:
        for c in r:
            m = re.search(r"с\s*\d{2}\.(\d{2})\.(\d{4})", str(c or ""))
            if m:
                return f"{m.group(2)}-{m.group(1)}"
    return None

# ── парсинг листов ──────────────────────────────────────────────────────────
def parse_sale(rows):
    """-> (rev_total, cost_total, items[])  берём строку «Итого» + позиции."""
    hi = find_header(rows, "блюдо")
    if hi < 0: return 0.0, 0.0, []
    hdr = [norm(c) for c in rows[hi]]
    def ci(*names):
        for nm in names:
            for j, c in enumerate(hdr):
                if nm in c: return j
        return -1
    c_name = ci("блюдо"); c_qty = ci("количество блюд")
    c_rev = ci("сумма со скидкой"); c_cost = ci("себестоимость, р")
    c_g1 = ci("группа блюда 1"); c_g2 = ci("группа блюда 2")
    items = []; rev_itogo = cost_itogo = None; rev_sum = cost_sum = 0.0
    for r in rows[hi+1:]:
        first = norm(r[c_name]) if 0 <= c_name < len(r) else ""
        if first == "":
            continue
        if first == "итого":
            rev_itogo  = num(r[c_rev])  if 0 <= c_rev  < len(r) else None
            cost_itogo = num(r[c_cost]) if 0 <= c_cost < len(r) else None
            continue
        rev = num(r[c_rev]) if 0 <= c_rev < len(r) else None
        qty = num(r[c_qty]) if 0 <= c_qty < len(r) else None
        cost = num(r[c_cost]) if 0 <= c_cost < len(r) else None
        if rev is None and qty is None and cost is None:
            continue
        items.append({"name": str(r[c_name]).strip(),
                      "g1": (str(r[c_g1]).strip() if 0 <= c_g1 < len(r) and r[c_g1] else ""),
                      "g2": (str(r[c_g2]).strip() if 0 <= c_g2 < len(r) and r[c_g2] else ""),
                      "qty": qty, "rev": rev or 0, "cost": cost or 0})
        rev_sum += rev or 0; cost_sum += cost or 0
    # «Итого» приоритетнее (точные значения iiko); иначе — сумма позиций
    rev_total  = rev_itogo  if rev_itogo  is not None else round(rev_sum, 2)
    cost_total = cost_itogo if cost_itogo is not None else round(cost_sum, 2)
    return rev_total, cost_total, items

def parse_writeoff(rows):
    """сумма колонки «Сумма, р.» по строкам-датам."""
    c_sum = 4
    for r in rows:
        for j, c in enumerate(r):
            if norm(c) == "сумма, р.":
                c_sum = j; break
    total = 0.0
    for r in rows:
        first = str(r[0]).strip() if r and r[0] is not None else ""
        if DATE_RE.match(first) and c_sum < len(r):
            v = num(r[c_sum])
            if v: total += v
    return round(total, 2)

# ── единая сборка summary (все ключи, что читает fnb.html) ───────────────────
def build_summary(rev_kitchen=0, rev_bs=0, rev_ba=0, rev_bfast=0, rev_banq=0, bp=None, bc=None,
                  cogs_kitchen=0, cogs_bs=0, cogs_ba=0, cogs_bfast=0, cogs_banq=0,
                  wo=None, other_wo=0.0, other_cost=0.0,
                  income_override=None, cost_override=None, fc_total_override=None,
                  banquet_hall=None, banquet_cork=None, services=None,
                  inv=None, guests_a=None, guests_b=None, extras=None):
    extras = extras or {}; wo = dict(wo or {}); inv = inv or {}
    if banquet_hall is None: banquet_hall = extras.get("rev_banquet_hall")
    if banquet_cork is None: banquet_cork = extras.get("rev_banquet_cork")
    if services is None:     services = extras.get("rev_services_total")
    ga = guests_a if guests_a is not None else extras.get("guests_alacarte")
    gb = guests_b if guests_b is not None else extras.get("guests_banquet")
    inv_total = inv.get("total") if inv.get("total") is not None else extras.get("inv_total")

    rev_alacarte_total = rev_kitchen + rev_bs + rev_ba
    cogs_alacarte = cogs_kitchen + cogs_bs + cogs_ba
    wo_sum = sum(v for v in wo.values() if v) + (other_wo or 0)
    income_total = income_override if income_override is not None else \
        (rev_kitchen + rev_bs + rev_ba + rev_bfast + rev_banq + (banquet_hall or 0) + (banquet_cork or 0) + (services or 0))
    cost_total = cost_override if cost_override is not None else \
        (cogs_bfast + cogs_alacarte + cogs_banq + wo_sum + (other_cost or 0))
    pct = lambda a, b: round(a / b * 100, 2) if b else None
    banquet_total = round(rev_banq + (banquet_hall or 0) + (banquet_cork or 0), 2) \
        if (banquet_hall is not None or banquet_cork is not None) else None
    R = lambda x: round(x, 2) if isinstance(x, (int, float)) else x
    return {
        "rev_breakfast_packets": R(bp) if bp is not None else None,
        "rev_breakfast_counter": R(bc) if bc is not None else None,
        "rev_breakfast": R(rev_bfast), "rev_kitchen": R(rev_kitchen),
        "rev_bar_soft": R(rev_bs), "rev_bar_alco": R(rev_ba),
        "rev_alacarte_total": R(rev_alacarte_total), "rev_banquet_food": R(rev_banq),
        "rev_banquet_hall": banquet_hall, "rev_banquet_cork": banquet_cork,
        "rev_banquet_total": banquet_total, "rev_services_total": services,
        "income_total": R(income_total),
        "cogs_breakfast": R(cogs_bfast), "cogs_alacarte": R(cogs_alacarte), "cogs_banquet": R(cogs_banq),
        "wo_comp": R(wo.get("wo_comp", 0)), "wo_spoilage": R(wo.get("wo_spoilage", 0)),
        "wo_represent": R(wo.get("wo_represent", 0)), "wo_deletions": R(wo.get("wo_deletions", 0)),
        "wo_staff": R(wo.get("wo_staff", 0)), "wo_free": R(wo.get("wo_free", 0.0)), "wo_dev": R(wo.get("wo_dev", 0)),
        "cost_total": R(cost_total),
        "inv_kitchen_start": inv.get("kitchen_start"), "inv_kitchen_buy": inv.get("kitchen_buy"),
        "inv_kitchen_end": inv.get("kitchen_end"), "inv_bar_start": inv.get("bar_start"),
        "inv_bar_buy": inv.get("bar_buy"), "inv_bar_end": inv.get("bar_end"), "inv_total": inv_total,
        "guests_alacarte": ga, "guests_banquet": gb,
        "check_alacarte": round(rev_alacarte_total / ga, 2) if ga else None,
        "check_banquet": round(rev_banq / gb, 2) if gb else None,
        "fc_kitchen": pct(cogs_kitchen, rev_kitchen), "fc_bar": pct(cogs_bs + cogs_ba, rev_bs + rev_ba),
        "fc_alacarte": pct(cogs_alacarte, rev_alacarte_total), "fc_breakfast": pct(cogs_bfast, rev_bfast),
        "fc_banquet": pct(cogs_banq, rev_banq),
        "fc_total": fc_total_override if fc_total_override is not None
                    else pct(cogs_bfast + cogs_alacarte + cogs_banq, income_total),
    }

# ── формат iiko (январь): 12 вкладок-категорий ──────────────────────────────
def parse_iiko(wb, mk, extras, log):
    cat_rev, cat_cost, sales = {}, {}, {}
    other_rev = other_cost = other_wo = 0.0
    wo = {"wo_deletions":0.0,"wo_comp":0.0,"wo_represent":0.0,"wo_staff":0.0,"wo_spoilage":0.0,"wo_dev":0.0}
    for sheet in wb.sheetnames:
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
        if not rows: continue
        kind, cat = classify(sheet, rows)
        if kind == "sale":
            rev, cost, items = parse_sale(rows)
            if cat:
                cat_rev[cat] = cat_rev.get(cat, 0) + rev
                cat_cost[cat] = cat_cost.get(cat, 0) + cost
                sales.setdefault(cat, []).extend(items)
            else:
                other_rev += rev; other_cost += cost
                log.append(f"  [{mk}] ⚠️ незнакомая ВКЛАДКА-ПРОДАЖИ «{sheet}» ({rev:,.0f} ₽) — в прочее")
        elif kind == "writeoff":
            total = parse_writeoff(rows)
            if cat is None: cat = cat_by_content(rows, WO_CATS)
            if cat: wo[cat] = wo.get(cat, 0) + total
            else:
                other_wo += total
                log.append(f"  [{mk}] ⚠️ незнакомая ВКЛАДКА-СПИСАНИЯ «{sheet}» ({total:,.0f} ₽) — в прочее")
    if not cat_rev and not any(wo.values()) and other_wo == 0 and other_rev == 0:
        return None
    g = lambda k: cat_rev.get(k, 0); c = lambda k: cat_cost.get(k, 0)
    bp = bc = 0.0
    for it in sales.get("breakfast", []):
        if "юл" in norm(it["g1"]) or (it.get("qty") or 0) >= 100: bp += it["rev"]
        else: bc += it["rev"]
    wo["wo_free"] = cat_cost.get("free", 0.0)   # «Без оплаты» — приход 0, важна себестоимость
    summary = build_summary(
        rev_kitchen=g("kitchen"), rev_bs=g("bar_soft"), rev_ba=g("bar_alco"),
        rev_bfast=g("breakfast"), rev_banq=g("banquet"), bp=bp, bc=bc,
        cogs_kitchen=c("kitchen"), cogs_bs=c("bar_soft"), cogs_ba=c("bar_alco"),
        cogs_bfast=c("breakfast"), cogs_banq=c("banquet"),
        wo=wo, other_wo=other_wo, other_cost=other_cost + cat_cost.get("free", 0.0), extras=extras)
    return {"summary": summary, "sales": sales, "banquets": []}

# ── формат A1-сводка (февраль): значения по меткам строк ─────────────────────
def a1_val(rows, *needles):
    ndl = [n.lower() for n in needles]
    for r in rows:
        j = norm(" | ".join("" if c is None else str(c) for c in r))
        if all(n in j for n in ndl):
            for k in range(2, len(r)):
                v = num(r[k])
                if v is not None: return v
    return None

def parse_a1(rows, extras, log):
    V = lambda *n: a1_val(rows, *n)
    rev_bfast = V("завтрак") or 0                     # строка-итог «Завтрак»
    bp = V("завтраки пакетные"); bc = V("завтраки от стойки")
    rev_kitchen = V("кухня") or 0
    rev_bs = V("бар напитки") or 0
    rev_ba = (V("бар алкоголь") or 0) + (V("бар пиво") or 0)
    rev_banq = V("питание") or 0                      # банкет «Питание» (первое вхождение)
    cogs_bfast = V("с/с", "завтрак") or 0
    cogs_kitchen = V("с/с", "а-ля карт") or 0
    cogs_bs = V("с/с", "напитки") or 0
    cogs_ba = (V("с/с", "пиво") or 0) + (V("с/с", "алкоголь") or 0)
    wo = {"wo_comp": V("комплемент") or 0, "wo_spoilage": V("порча") or 0,
          "wo_represent": V("представительск") or 0, "wo_staff": V("питание сотрудник") or 0,
          "wo_dev": V("проработка") or 0, "wo_deletions": V("удаление") or 0}
    inv = {"kitchen_start": V("кухня начало"), "kitchen_buy": V("кухня закуп"), "kitchen_end": V("кухня конец"),
           "bar_start": V("бар начало"), "bar_buy": V("бар закуп"), "bar_end": V("бар конец"),
           "total": V("итого товарный остаток")}
    return {"summary": build_summary(
        rev_kitchen=rev_kitchen, rev_bs=rev_bs, rev_ba=rev_ba, rev_bfast=rev_bfast, rev_banq=rev_banq,
        bp=bp, bc=bc, cogs_kitchen=cogs_kitchen, cogs_bs=cogs_bs, cogs_ba=cogs_ba,
        cogs_bfast=cogs_bfast, cogs_banq=(V("с/с", "банкет") or 0), wo=wo,
        income_override=(V("доход fb") or V("итого приход дс")),
        cost_override=V("итого расход продуктов"), fc_total_override=V("общий fc службы"),
        banquet_hall=V("аренда зала"), banquet_cork=V("пробковый"), services=V("услуги"),
        inv=inv, guests_a=V("количество гостей по меню"), guests_b=V("количество гостей на банкет"),
        extras=extras), "sales": {}, "banquets": []}

# ── формат Sheet1 (март–июнь): категории × колонка «Общее» ───────────────────
WO_WORDS = {"представительск": "wo_represent", "порча": "wo_spoilage", "комплимент": "wo_comp",
            "удаление": "wo_deletions", "проработ": "wo_dev", "стафф": "wo_staff"}

def parse_sheet1(rows, extras, log):
    col_total = None
    for r in rows:                                    # колонка «Общее» из заголовка
        for j, cc in enumerate(r):
            if norm(cc) == "общее": col_total = j; break
        if col_total is not None: break
    if col_total is None:
        col_total = max((len(r) for r in rows), default=1) - 1
    rev = {}; cogs = {}; income = None; other_wo = 0.0
    wo = {v: 0.0 for v in set(WO_WORDS.values())}
    section = "income"
    for r in rows:
        label = norm(r[2]) if len(r) > 2 and r[2] else ""
        if not label: continue
        if label == "себес": section = "cogs"; continue
        if label.startswith("кост"): break
        v = num(r[col_total]) if col_total < len(r) else None
        if v is None: continue
        if section == "income":
            if "доход fb" in label: income = v
            elif "напитки" in label: rev["bar_soft"] = rev.get("bar_soft", 0) + v
            elif label.startswith("пиво") or "крепкая" in label or "вино" in label or "коктел" in label:
                rev["bar_alco"] = rev.get("bar_alco", 0) + v
            elif "банкет" in label: rev["banquet"] = rev.get("banquet", 0) + v
            elif "меню аля" in label or "аля карт" in label: rev["kitchen"] = rev.get("kitchen", 0) + v
            elif "завтрак" in label: rev["breakfast"] = rev.get("breakfast", 0) + v
            elif "услуг" in label or "рум сервис" in label: rev["services"] = rev.get("services", 0) + v
            elif "пробковый" in label: rev["cork"] = rev.get("cork", 0) + v
        else:                                          # секция себес: себес-категории + списания
            wcat = next((k for w, k in WO_WORDS.items() if w in label), None)
            if wcat: wo[wcat] += v
            elif "маркетинг" in label: other_wo += v
            elif "напитки" in label: cogs["bar_soft"] = cogs.get("bar_soft", 0) + v
            elif label.startswith("пиво") or "крепкая" in label or "вино" in label:
                cogs["bar_alco"] = cogs.get("bar_alco", 0) + v
            elif "банкет" in label: cogs["banquet"] = cogs.get("banquet", 0) + v
            elif "меню аля" in label or "аля карт" in label: cogs["kitchen"] = cogs.get("kitchen", 0) + v
            elif "завтрак" in label: cogs["breakfast"] = cogs.get("breakfast", 0) + v
    return {"summary": build_summary(
        rev_kitchen=rev.get("kitchen", 0), rev_bs=rev.get("bar_soft", 0), rev_ba=rev.get("bar_alco", 0),
        rev_bfast=rev.get("breakfast", 0), rev_banq=rev.get("banquet", 0),
        cogs_kitchen=cogs.get("kitchen", 0), cogs_bs=cogs.get("bar_soft", 0), cogs_ba=cogs.get("bar_alco", 0),
        cogs_bfast=cogs.get("breakfast", 0), cogs_banq=cogs.get("banquet", 0),
        wo=wo, other_wo=other_wo, income_override=income,
        banquet_cork=rev.get("cork"), services=rev.get("services"), extras=extras),
        "sales": {}, "banquets": []}

# ── определить формат книги ─────────────────────────────────────────────────
def detect_format(wb):
    for sh in wb.sheetnames:
        rows = [list(r) for r in wb[sh].iter_rows(values_only=True, max_row=45)]
        txt = norm(" | ".join("" if c is None else str(c) for r in rows for c in r))
        if "блюдо" in txt and "сумма со скидкой" in txt: return "iiko"
        if "приход спин" in txt or "сводный отчет по движению" in txt or "итого приход дс" in txt: return "a1"
        if "доход fb" in txt: return "sheet1"
    return None

# ── сборка одного месяца (скачать -> определить формат -> распарсить) ─────────
def build_month(entry, log):
    mk, fid, extras = entry["month"], entry["file_id"], entry["extras"]
    data = None
    for u in (f"https://docs.google.com/spreadsheets/d/{fid}/export?format=xlsx",
              f"https://drive.google.com/uc?export=download&id={fid}"):
        try:
            raw = http_get(u)
            if raw[:2] == b"PK": data = raw; break
        except Exception as e:
            log.append(f"  [{mk}] download err: {e}")
    if data is None:
        log.append(f"  [{mk}] ❌ не удалось скачать файл {fid}")
        return None
    xpath = RAW / f"{mk}.xlsx"; xpath.write_bytes(data)
    wb = openpyxl.load_workbook(xpath, read_only=True, data_only=True)

    fmt = detect_format(wb)
    if fmt is None:
        log.append(f"  [{mk}] ⛔ формат не распознан (вкладки: {wb.sheetnames}) — месяц ПРОПУЩЕН")
        return None
    if fmt == "iiko":
        res = parse_iiko(wb, mk, extras, log)
    else:
        # A1 / Sheet1 — данные на одной вкладке; берём первую непустую
        rows = []
        for sh in wb.sheetnames:
            rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
            rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
            if rows: break
        res = parse_a1(rows, extras, log) if fmt == "a1" else parse_sheet1(rows, extras, log)

    if res is None:
        log.append(f"  [{mk}] ⛔ формат «{fmt}» дал пусто — месяц ПРОПУЩЕН")
        return None
    s = res["summary"]
    wo_sum = sum(v for v in [s["wo_comp"], s["wo_spoilage"], s["wo_represent"], s["wo_deletions"],
                             s["wo_staff"], s["wo_dev"], s.get("wo_free", 0)] if v)
    log.append(f"  [{mk}] ✓ [{fmt}] выручка {s['income_total']:,.0f} ₽ | списания {wo_sum:,.0f} | FC {s['fc_total']}%")
    return res

# ── main ────────────────────────────────────────────────────────────────────
def build():
    log = []
    reg = fetch_registry()
    log.append(f"Реестр: {len(reg)} месяц(ев) со ссылками: " + ", ".join(e["month"] for e in reg))
    months = {}
    for entry in reg:
        m = build_month(entry, log)
        if m: months[entry["month"]] = m
    data = {"generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), "months": months}
    js = "// fnb_data.js — auto-generated (registry+all-tabs), do not edit\nwindow.UTRAU_FNB=" + \
         json.dumps(data, ensure_ascii=False) + ";"
    OUT.write_text(js, encoding="utf-8")
    print("\n".join(log))
    print(f"\nGenerated {OUT} ({len(months)} month(s))")

if __name__ == "__main__":
    build()
