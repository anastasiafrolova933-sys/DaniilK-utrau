# -*- coding: utf-8 -*-
"""
build_spa.py — СПА / Банный комплекс Утрау (доп. услуги ExtraServices).
Скачивает весь workbook (XLSX) и парсит помесячные вкладки (услуга -> выручка)
→ data/spa_data.js  (window.UTRAU_SPA)

Источник: Google Sheets 1eMzwZszovZ2AqIpaMS6PD-tvJvYdGZDD5DEaj8Qhu0w
XLSX-подход: новые месяцы (вкладки) подхватываются автоматически, gid'ы не нужны.
Запуск: python scripts/build_spa.py
"""
import re, io, json, urllib.request, datetime
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "spa_raw.xlsx"
OUT  = ROOT / "data" / "spa_data.js"
SHEET_ID = "1eMzwZszovZ2AqIpaMS6PD-tvJvYdGZDD5DEaj8Qhu0w"

MONTHS_RU = {"янв":1,"фев":2,"мар":3,"апр":4,"май":5,"июн":6,
             "июл":7,"авг":8,"сен":9,"окт":10,"ноя":11,"дек":12}

# Категории по ключевым словам (порядок важен: сначала специфичные)
def categorize(name):
    s = name.lower()
    if "джакузи" in s: return "Джакузи"
    if "массаж" in s: return "Массаж"
    if any(k in s for k in ["парени","парение","веник","прогрев","щабын","скраб","обертыв","релакс"]): return "Парение и уход"
    if any(k in s for k in ["баня","банного комплекса","банный","бк ","банного","термальн","новогодний утрау"]): return "Баня / Банный комплекс"
    if any(k in s for k in ["сауна","бассейн","подогрев"]): return "Сауна и бассейн"
    if any(k in s for k in ["тариф","тапочки","прокат","халат"]): return "Тарифы и прочее"
    return "Прочее"

def fetch_xlsx():
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        XLSX.write_bytes(r.read())
    print(f"  XLSX downloaded ({XLSX.stat().st_size//1024} KB)")

def detect_month(name):
    """Имя вкладки -> 'YYYY-MM' или None."""
    m = re.search(r'_\d{1,2}\.(\d{2})\.(\d{4})', name)   # суффикс даты _ДД.ММ.ГГГГ (2024)
    if m:
        return f"{m.group(2)}-{m.group(1)}"
    low = name.lower()
    mon = None
    for k, v in MONTHS_RU.items():
        if k in low:
            mon = v; break
    year = None
    ym = re.search(r'20(\d{2})', name)          # 2025 / 2026
    if ym:
        year = int("20" + ym.group(1))
    else:
        y2 = re.search(r'(\d{2})(?!\d)', name)   # хвостовые 2 цифры: Янв25 -> 25
        if y2: year = 2000 + int(y2.group(1))
    if mon and year:
        return f"{year}-{mon:02d}"
    return None

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return round(float(v), 2)
    s = str(v).strip().replace("\xa0","").replace(" ","").replace("₽","").replace(",",".")
    try: return round(float(s), 2)
    except ValueError: return None

def parse_sheet(ws):
    services = []
    for row in ws.iter_rows(values_only=True):
        if not row: continue
        name = row[0]
        if name is None: continue
        name = str(name).strip()
        if not name or name.lower() == "название": continue
        rev = num(row[1]) if len(row) > 1 else None
        if rev is None: continue
        services.append({"name": name, "rev": rev, "cat": categorize(name)})
    return services

def build():
    fetch_xlsx()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    months = {}
    skipped = []
    for sheet in wb.sheetnames:
        mo = detect_month(sheet)
        ws = wb[sheet]
        services = parse_sheet(ws)
        if not services:
            skipped.append((sheet, "пусто"))
            continue
        if not mo:
            skipped.append((sheet, "не распознан месяц"))
            continue
        total = round(sum(s["rev"] for s in services), 2)
        # если месяц уже есть (дубль) — берём больший по сумме
        if mo not in months or total > months[mo]["total"]:
            months[mo] = {"total": total, "services": services}
    data = {"generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), "months": months}
    js = "// spa_data.js — auto-generated, do not edit\nwindow.UTRAU_SPA=" + \
         json.dumps(data, ensure_ascii=False) + ";"
    OUT.write_text(js, encoding="utf-8")
    print(f"\nМесяцев: {len(months)}")
    for mo in sorted(months):
        print(f"  {mo}: {months[mo]['total']:>14,.0f}  ({len(months[mo]['services'])} услуг)")
    if skipped:
        print("Пропущены:", ", ".join(f"{s}({r})" for s,r in skipped))
    print(f"Generated {OUT}")

if __name__ == "__main__":
    build()
